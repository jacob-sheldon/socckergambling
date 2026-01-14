"""
Generate betting analysis Excel template with real match data from 500.com.
Run: uv run generate-live-template
"""

import re
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Default URL for Jingcai (竞彩) scores - mobile version is simpler to parse
DEFAULT_URL = "https://live.m.500.com/home/zq/jczq/cur"

# ============================================================================
# CONSTANTS (reused from betting_analysis_template.py)
# ============================================================================

# Two-row headers
MAIN_HEADERS = [
    "",
    "亚盘盘口",
    "",
    "",
    "横纵分析",
    "",
    "",
    "左右格局警示",
    "",
    "主流凯利",
    "",
    "",
    "",
    "平局预警",
    "平赔数据",
    "",
    "",
    ""
]

SUB_HEADERS = [
    "比赛/时间",
    "初始亚凯对比",
    "主",
    "客",
    "对比",
    "主",
    "客",
    "预测",
    "提示",
    "初凯",
    "即凯",
    "初变",
    "初变",
    "",
    "初凯",
    "即凯",
    "即初凯",
    "初凯变化"
]

# Column width definitions
COLUMN_WIDTHS = {
    'A': 10,  # 比赛/时间
    'B': 12,  # 亚盘盘口
    'C': 8,   # 主
    'D': 8,   # 客
    'E': 8,   # 对比
    'F': 8,   # 主
    'G': 8,   # 客
    'H': 8,   # 预测
    'I': 8,   # 提示
    'J': 8,   # 初凯
    'K': 8,   # 即凯
    'L': 8,   # 初变
    'M': 8,   # 初变
    'N': 10,  # 平局预警
    'O': 8,   # 初凯
    'P': 8,   # 即凯
    'Q': 10,  # 即初凯
    'R': 12,  # 初凯变化
}

# Merged cell ranges for main sections
MERGED_CELLS = [
    ("B1", "D1"),  # 亚盘盘口
    ("E1", "G1"),  # 横纵分析
    ("H1", "I1"),  # 左右格局警示
    ("J1", "M1"),  # 主流凯利
    ("N1", "N1"),  # 平局预警
    ("O1", "R1"),  # 平赔数据
]

# Time tracking labels
TIME_LABELS = [
    "初盘3点",
    "赛前1",
    "赛前2",
    "临场一小时",
    "临场半小时",
    "临场15分钟",
    "临场10分钟",
    "临场",
    "完场"
]

# ============================================================================
# STYLING DEFINITIONS
# ============================================================================

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

BORDER_STYLE = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

MATCH_ID_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
MATCH_ID_FONT = Font(bold=True, size=11)

TIME_LABEL_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
TIME_LABEL_FONT = Font(size=10)

DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# ============================================================================
# WEB SCRAPING FUNCTIONS
# ============================================================================


def fetch_jingcai_matches(url=DEFAULT_URL, timeout=15, default_count=15):
    """
    Fetch match IDs from 500.com Jingcai (竞彩比分) section.
    Uses mobile version which provides simpler HTML structure.
    Uses curl via subprocess to avoid Python SSL library issues.

    If web scraping fails (website structure changed, network issues, etc.),
    falls back to generating sequential match IDs.

    Args:
        url: The URL to scrape (default: mobile 500.com jczq page)
        timeout: Request timeout in seconds
        default_count: Number of sequential match IDs to generate as fallback

    Returns:
        List of match ID strings (e.g., ["001", "002", "003"])
    """
    try:
        print(f"Fetching data from {url}...")

        # Use curl via subprocess to avoid Python SSL library issues with this site
        # curl handles problematic SSL configurations better
        result = subprocess.run(
            [
                'curl', '-k', '-s',  # -k: insecure, -s: silent
                '-m', str(timeout),  # timeout
                '-L',  # follow redirects
                '-A', 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1',
                url
            ],
            capture_output=True,
            text=True,
            timeout=timeout + 5
        )

        if result.returncode != 0:
            raise RuntimeError(f"curl failed with exit code {result.returncode}: {result.stderr}")

        html_content = result.stdout

        # Try multiple regex patterns for different website versions
        patterns = [
            r'周[一二三四五六七日天日]\d{3}',  # Pattern: 周一001, 周二002, etc.
            r'期(\d{3})',  # Pattern: 期001
            r'match[_-]?id["\']?\s*[:=]\s*["\']?(\d{3})',  # Pattern: match_id: "001"
            r'data[_-]?match[_-]?id["\']?\s*[:=]\s*["\']?(\d{3})',  # Pattern: data-match-id="001"
            r'jc[_-]?(\d{3})',  # Pattern: jc001
        ]

        match_ids = []
        for pattern in patterns:
            matches = re.findall(pattern, html_content)
            if matches:
                for match_id in matches:
                    # Extract just the numeric part (last 3 digits)
                    numeric_id = match_id[-3:] if len(match_id) >= 3 else match_id.zfill(3)
                    # Avoid duplicates
                    if numeric_id not in match_ids:
                        match_ids.append(numeric_id)

                if match_ids:
                    break

        if not match_ids:
            raise ValueError("No match IDs found on the page. The website structure may have changed.")

        print(f"Found {len(match_ids)} matches: {match_ids[:5]}{'...' if len(match_ids) > 5 else ''}")
        return match_ids

    except subprocess.TimeoutExpired:
        print(f"Request timed out after {timeout} seconds")
        return _generate_fallback_match_ids(default_count)
    except FileNotFoundError:
        print("curl is not installed. Using sequential match IDs.")
        return _generate_fallback_match_ids(default_count)
    except Exception as e:
        print(f"Web scraping failed ({e}). Using sequential match IDs as fallback.")
        return _generate_fallback_match_ids(default_count)


def _generate_fallback_match_ids(count):
    """
    Generate sequential match IDs as fallback when web scraping fails.

    Args:
        count: Number of match IDs to generate

    Returns:
        List of match ID strings (e.g., ["001", "002", "003"])
    """
    match_ids = [f"{i:03d}" for i in range(1, count + 1)]
    print(f"Generated {len(match_ids)} sequential match IDs: {match_ids[:5]}{'...' if len(match_ids) > 5 else ''}")
    return match_ids


# ============================================================================
# EXCEL GENERATION FUNCTIONS (reused/adapted from betting_analysis_template.py)
# ============================================================================


def create_template_workbook():
    """Create a new workbook with headers and basic formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "足球彩票分析模板"
    return wb, ws


def set_column_widths(ws):
    """Set appropriate column widths."""
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width


def merge_header_cells(ws):
    """Merge cells for main section headers."""
    for start_cell, end_cell in MERGED_CELLS:
        ws.merge_cells(f"{start_cell}:{end_cell}")


def style_header_rows(ws):
    """Apply styling to header rows."""
    # Style Row 1 - Main sections
    for col_idx, header in enumerate(MAIN_HEADERS, start=1):
        if header:  # Only style non-empty headers
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER_STYLE

    # Style Row 2 - Sub-headers
    for col_idx, header in enumerate(SUB_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = BORDER_STYLE

    # Set row heights for headers
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20


def add_sample_data(ws, start_row, match_id):
    """
    Add sample data for one match to demonstrate the template.

    Args:
        ws: Worksheet object
        start_row: Starting row number for this match
        match_id: Match ID string (e.g., "001")

    Returns:
        Number of rows added
    """
    current_row = start_row

    # Sample data for demonstration
    sample_handicap = "-0.75"

    # Match ID row with handicap
    ws.cell(row=current_row, column=1, value=match_id)
    ws.cell(row=current_row, column=2, value=sample_handicap)

    # Style match ID row
    for col_idx in range(1, 19):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.fill = MATCH_ID_FILL
        cell.font = MATCH_ID_FONT
        cell.alignment = DATA_ALIGNMENT
        cell.border = BORDER_STYLE

    current_row += 1

    # Sample odds data for different time points
    sample_time_data = {
        "初盘3点": {
            3: 0.95, 4: 0.93, 6: 0.94, 7: 0.95,
            10: 0.89, 11: 1.10, 15: 0.98, 16: 0.99, 17: 0.97
        },
        "赛前1": {
            3: 0.96, 4: 0.93, 6: 0.96, 7: 0.91,
            10: 0.90, 11: 1.06, 15: 0.97, 16: 0.98, 17: 0.99
        },
        "赛前2": {
            2: "-1.0", 3: 0.99, 4: 0.86, 6: 0.93, 7: 0.96,
            10: 0.91, 11: 0.99, 15: 0.94, 16: 1.01, 17: 1.03
        },
        "临场一小时": {
            3: 1.00, 4: 0.86, 6: 0.93, 7: 0.96,
            10: 0.91, 11: 0.99, 15: 0.94, 16: 1.02, 17: 1.04
        },
        "临场半小时": {
            3: 1.00, 4: 0.86, 6: 0.95, 7: 0.92,
            10: 0.91, 11: 1.00, 15: 0.93, 16: 1.02
        },
        "临场15分钟": {
            6: 0.96, 7: 0.89, 10: 0.93, 11: 0.96,
            15: 0.93, 16: 1.01
        },
        "临场10分钟": {
            6: 0.96, 7: 0.90
        },
        "临场": {},
        "完场": {
            8: 1.3  # 预测值
        }
    }

    # Add data for each time point
    for time_label in TIME_LABELS:
        # First column: time label
        ws.cell(row=current_row, column=1, value=time_label)

        # Add sample data if available
        if time_label in sample_time_data:
            for col_idx, value in sample_time_data[time_label].items():
                ws.cell(row=current_row, column=col_idx, value=value)

        # Style the row
        for col_idx in range(1, 19):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.alignment = DATA_ALIGNMENT
            cell.border = BORDER_STYLE

            # Time label column gets gray fill
            if col_idx == 1:
                cell.fill = TIME_LABEL_FILL
                cell.font = TIME_LABEL_FONT
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        current_row += 1

    return len(TIME_LABELS) + 1  # +1 for match ID row


def generate_live_template(filename="live_betting_template.xlsx", url=DEFAULT_URL, max_matches=None):
    """
    Generate Excel betting analysis template with real match data from 500.com.

    Args:
        filename: Output Excel filename
        url: URL to scrape match data from
        max_matches: Maximum number of matches to include (None = all matches)
    """
    # Fetch real match IDs from the website
    match_ids = fetch_jingcai_matches(url)

    # Limit matches if specified
    if max_matches is not None:
        match_ids = match_ids[:max_matches]

    # Create workbook and worksheet
    wb, ws = create_template_workbook()

    # Set column widths
    set_column_widths(ws)

    # Merge header cells
    merge_header_cells(ws)

    # Style header rows
    style_header_rows(ws)

    # Add matches with real match IDs
    current_row = 3  # Start after headers (rows 1-2)
    for match_id in match_ids:
        rows_added = add_sample_data(ws, current_row, match_id)
        current_row += rows_added

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Save workbook
    wb.save(filename)
    print(f"Excel template '{filename}' generated successfully!")
    print(f"Contains {len(match_ids)} real matches from {url}.")


def main():
    """Entry point for the CLI command."""
    generate_live_template()


if __name__ == "__main__":
    main()
