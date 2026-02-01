"""
Generate betting analysis Excel template with comprehensive real match data from live.500.com.
This uses Playwright to extract the full table data from the Jingcai score page.

Run: uv run generate-browser-template
"""

import asyncio
import json
import os
import re
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


LAST_SCRAPE_ERROR = ""


def _get_default_playwright_cache_dir() -> Path:
    """Return the default Playwright browser cache directory for this platform."""
    if sys.platform.startswith("win"):
        base = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local"))
        return base / "ms-playwright"
    if sys.platform == "darwin":
        return Path.home() / "Library" / "Caches" / "ms-playwright"
    return Path.home() / ".cache" / "ms-playwright"


def _setup_bundled_browser():
    """
    Configure Playwright to use the bundled browser in the .app package.
    This must be called before importing Playwright modules.
    """
    # Check if we're running from a PyInstaller bundle
    if not getattr(sys, 'frozen', False):
        return False

    # Try multiple strategies to find the bundled browsers directory
    browsers_path = None

    # Strategy 1: Check sys._MEIPASS (works for onedir builds)
    if hasattr(sys, '_MEIPASS'):
        path = Path(sys._MEIPASS) / 'ms-playwright'
        if path.exists():
            browsers_path = path
            print(f"[DEBUG] Found browsers via _MEIPASS: {browsers_path}")

    # Strategy 2: For macOS .app bundles, find Resources via executable path
    # .app structure: AppName.app/Contents/MacOS/executable -> ../Resources
    if browsers_path is None and 'MacOS' in Path(sys.executable).parts:
        exec_path = Path(sys.executable)
        resources_path = exec_path.parent.parent / 'Resources' / 'ms-playwright'
        if resources_path.exists():
            browsers_path = resources_path
            print(f"[DEBUG] Found browsers via executable path: {browsers_path}")

    # Strategy 3: Try looking relative to the executable (for onefile builds)
    if browsers_path is None:
        exec_dir = Path(sys.executable).parent
        path = exec_dir / 'ms-playwright'
        if path.exists():
            browsers_path = path
            print(f"[DEBUG] Found browsers via exec_dir: {browsers_path}")

    if browsers_path:
        os.environ['PLAYWRIGHT_BROWSERS_PATH'] = str(browsers_path)  # Directory containing chromium-*
        print(f"[DEBUG] Set PLAYWRIGHT_BROWSERS_PATH to: {browsers_path}")
        return True

    # Fall back to system cache if browsers were installed separately.
    cache_dir = _get_default_playwright_cache_dir()
    if cache_dir.exists():
        os.environ.setdefault('PLAYWRIGHT_BROWSERS_PATH', str(cache_dir))
        print(f"[DEBUG] Using Playwright cache from system: {cache_dir}")
        return True

    print(f"[DEBUG] Could not find bundled browsers. sys.frozen={sys.frozen}, _MEIPASS={getattr(sys, '_MEIPASS', 'N/A')}, executable={sys.executable}")
    return False


# Setup bundled browser path before importing Playwright
_setup_bundled_browser()

# URLs for different data sources
BASE_URL = "https://live.500.com/"  # Desktop version with full table data
DESKTOP_URL = "https://www.500.com/jczq/"
ODDS_URL = "https://odds.500.com/"

# ============================================================================
# DATA STRUCTURES
# ============================================================================


@dataclass
class MatchData:
    """Comprehensive match data structure from live.500.com table."""
    match_id: str           # 场次: 周一001, 周二002, etc.
    league: str             # 赛事: 澳超, 非洲杯, etc.
    round: str              # 轮次: 第17轮, 1/8决赛, etc.
    match_time: str         # 比赛时间: 01-05 16:00
    status: str             # 状态: 未, 进行中, 完场
    home_team: str          # 主队: [04]麦克阿瑟FC
    home_rank: str          # 主队排名: 04
    handicap: str           # 让球: 受平手/半球, 一球/球半, etc.
    away_team: str          # 客队: 奥克兰FC[01]
    away_rank: str          # 客队排名: 01
    halftime_score: str     # 半场比分
    win_odds: str           # 胜负奖金
    let_odds: str           # 让球奖金
    avg_euro: str           # 平均欧赔
    william_odds: str       # 威廉赔率
    aust_odds: str          # 澳彩赔率
    bet365_odds: str        # 365赔率
    royal_odds: str         # 皇者赔率
    # Asian handicap analysis data (from "亚" link pages)
    asian_handicap: str = ""   # 盘 - handicap value from analysis page (e.g., "半球", "一球")
    home_water: str = ""        # 水1 - home team odds from Crown (冠) row
    away_water: str = ""        # 水2 - away team odds from Crown (冠) row
    analysis_url: str = ""      # URL of the Asian handicap analysis page
    euro_odds_url: str = ""     # URL of the European odds (百家欧赔) page
    # European odds (百家欧赔) - Kelly index from Crown (冠) row
    euro_kelly_win: str = ""    # 即时凯利 胜
    euro_kelly_draw: str = ""   # 即时凯利 平
    euro_kelly_lose: str = ""   # 即时凯利 负
    euro_kelly_win_2: str = ""  # 即时凯利 第二行 胜
    euro_kelly_draw_2: str = "" # 即时凯利 第二行 平
    euro_kelly_lose_2: str = "" # 即时凯利 第二行 负
    is_fallback: bool = False   # True if this is placeholder data from error fallback

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for Excel serialization."""
        return {
            'match_id': self.match_id,
            'league': self.league,
            'round': self.round,
            'match_time': self.match_time,
            'status': self.status,
            'home_team': self.home_team,
            'home_rank': self.home_rank,
            'handicap': self.handicap,
            'away_team': self.away_team,
            'away_rank': self.away_rank,
            'halftime_score': self.halftime_score,
            'win_odds': self.win_odds,
            'let_odds': self.let_odds,
            'avg_euro': self.avg_euro,
            'william_odds': self.william_odds,
            'aust_odds': self.aust_odds,
            'bet365_odds': self.bet365_odds,
            'royal_odds': self.royal_odds,
            'asian_handicap': self.asian_handicap,
            'home_water': self.home_water,
            'away_water': self.away_water,
            'analysis_url': self.analysis_url,
            'euro_odds_url': self.euro_odds_url,
            'euro_kelly_win': self.euro_kelly_win,
            'euro_kelly_draw': self.euro_kelly_draw,
            'euro_kelly_lose': self.euro_kelly_lose,
            'euro_kelly_win_2': self.euro_kelly_win_2,
            'euro_kelly_draw_2': self.euro_kelly_draw_2,
            'euro_kelly_lose_2': self.euro_kelly_lose_2,
        }


# ============================================================================
# EXCEL TEMPLATES CONSTANTS
# ============================================================================

MAIN_HEADERS = [
    "",
    "亚盘盘口",
    "",
    "",
    "横纵分析",
    "",
    "",
    "左右格局警示",
    "",
    "主流凯利",
    "",
    "",
    "",
    "平局预警",
    "平赔数据",
    "",
    "",
    ""
]

SUB_HEADERS = [
    "比赛/时间",
    "初始亚凯对比",
    "主",
    "客",
    "对比",
    "主",
    "客",
    "预测",
    "提示",
    "初凯",
    "即凯",
    "初变",
    "初变",
    "",
    "初凯",
    "即凯",
    "即初凯",
    "初凯变化"
]

COLUMN_WIDTHS = {
    'A': 10, 'B': 12, 'C': 8, 'D': 8, 'E': 8,
    'F': 8, 'G': 8, 'H': 8, 'I': 8, 'J': 8,
    'K': 8, 'L': 8, 'M': 8, 'N': 10, 'O': 8,
    'P': 8, 'Q': 10, 'R': 12
}

MERGED_CELLS = [
    ("B1", "D1"), ("E1", "G1"), ("H1", "I1"),
    ("J1", "M1"), ("N1", "N1"), ("O1", "R1")
]

TIME_LABELS = [
    "初盘3点", "赛前1", "赛前2", "临场一小时",
    "临场半小时", "临场15分钟", "临场10分钟", "临场", "完场"
]

# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

BORDER_STYLE = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

MATCH_ID_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
MATCH_ID_FONT = Font(bold=True, size=11)

TIME_LABEL_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
TIME_LABEL_FONT = Font(size=10)

DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")


# ============================================================================
# BROWSER AUTOMATION WITH PLAYWRIGHT
# ============================================================================


async def fetch_matches_with_browser(
    url: str = BASE_URL,
    headless: bool = True,
    timeout: int = 60000
) -> List[MatchData]:
    """
    Fetch comprehensive match data from live.500.com using Playwright.
    Extracts the full Jingcai betting table with all columns.

    Args:
        url: URL to scrape (default: live.500.com)
        headless: Run browser in headless mode (no GUI)
        timeout: Page load timeout in milliseconds

    Returns:
        List of MatchData objects with comprehensive match information
    """
    global LAST_SCRAPE_ERROR
    LAST_SCRAPE_ERROR = ""
    from playwright.async_api import async_playwright, Error as PlaywrightError

    print(f"正在从 {url} 获取竞彩比分数据...")
    print("=" * 80)

    try:
        async with async_playwright() as p:
            # Launch browser with anti-detection settings
            browser = await p.chromium.launch(
                headless=headless,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox',
                ]
            )

            # Create context with realistic user agent (desktop)
            context = await browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                viewport={'width': 1920, 'height': 1080},
                locale='zh-CN',
            )

            page = await context.new_page()

            # Navigate to URL and wait for network idle
            print("正在加载页面...")
            await page.goto(url, wait_until='domcontentloaded', timeout=timeout)

            # Wait for the table to load
            try:
                await page.wait_for_selector('table', timeout=10000)
                print("✓ 表格加载成功")
            except Exception:
                print("⚠ 未找到表格，继续尝试...")

            # Execute JavaScript to extract the full table data
            print("正在提取表格数据...")

            extraction_script = """
            () => {
                const matches = [];

                // Find the main table with Jingcai scores
                const tables = document.querySelectorAll('table');
                let targetTable = null;

                // Look for table with match rows
                for (const table of tables) {
                    const rows = table.querySelectorAll('tr');
                    if (rows.length > 2) {
                        const firstRow = rows[0];
                        if (firstRow.innerText.includes('场次') || firstRow.innerText.includes('赛事')) {
                            targetTable = table;
                            break;
                        }
                    }
                }

                if (!targetTable) {
                    console.log('未找到目标表格');
                    return matches;
                }

                console.log('找到表格，行数:', targetTable.querySelectorAll('tr').length);

                // Extract data from each row
                const rows = targetTable.querySelectorAll('tr');
                for (let i = 0; i < rows.length; i++) {
                    const row = rows[i];
                    const cells = row.querySelectorAll('td');

                    if (cells.length < 5) continue; // Skip header rows or empty rows

                    try {
                        // Extract match ID (e.g., "周一001")
                        const firstCell = cells[0].innerText.trim();
                        const matchIdMatch = firstCell.match(/周[一二三四五六七日天日](\\d{3})/);
                        if (!matchIdMatch) continue;

                        const match_id = firstCell;

                        // Extract league (赛事)
                        const league = cells[1] ? cells[1].innerText.trim() : '';

                        // Extract round (轮次)
                        const round = cells[2] ? cells[2].innerText.trim() : '';

                        // Extract match time
                        const match_time = cells[3] ? cells[3].innerText.trim() : '';

                        // Extract status
                        const status = cells[4] ? cells[4].innerText.trim() : '';

                        // Extract home team with ranking
                        let home_team_raw = cells[5] ? cells[5].innerText.trim() : '';
                        // Handle completed matches with score embedded
                        if (home_team_raw.includes('\\n')) {
                            const lines = home_team_raw.split('\\n');
                            home_team_raw = lines[0];
                        }
                        const home_rank_match = home_team_raw.match(/\\[(\\d+)\\]/);
                        const home_rank = home_rank_match ? home_rank_match[1] : '';
                        const home_team = home_team_raw.replace(/\\[\\d+\\]/, '').trim();

                        // Extract handicap (让球)
                        const handicap = cells[6] ? cells[6].innerText.trim() : '';

                        // Extract away team with ranking
                        let away_team_raw = cells[7] ? cells[7].innerText.trim() : '';
                        // Handle completed matches with score embedded
                        if (away_team_raw.includes('\\n')) {
                            const lines = away_team_raw.split('\\n');
                            away_team_raw = lines[0];
                        }
                        const away_rank_match = away_team_raw.match(/\\[(\\d+)\\]/);
                        const away_rank = away_rank_match ? away_rank_match[1] : '';
                        const away_team = away_team_raw.replace(/\\[\\d+\\]/, '').trim();

                        // Extract halftime score
                        let halftime_score = cells[8] ? cells[8].innerText.trim() : '';
                        // Clean up multiline scores
                        if (halftime_score.includes('\\n')) {
                            const lines = halftime_score.split('\\n');
                            halftime_score = lines[lines.length - 1].trim();
                        }

                        // Extract odds data
                        const win_odds = cells[9] ? cells[9].innerText.trim() : '';
                        const let_odds = cells[10] ? cells[10].innerText.trim() : '';
                        const avg_euro = cells[11] ? cells[11].innerText.trim() : '';

                        // Extract company odds
                        const william_odds = cells[12] ? cells[12].innerText.trim() : '';
                        const aust_odds = cells[13] ? cells[13].innerText.trim() : '';
                        const bet365_odds = cells[14] ? cells[14].innerText.trim() : '';
                        const royal_odds = cells[15] ? cells[15].innerText.trim() : '';

                        // Extract analysis links (亚盘对比/百家欧赔)
                        let analysis_url = '';
                        let euro_odds_url = '';
                        for (let j = 0; j < cells.length; j++) {
                            const links = cells[j].querySelectorAll('a');
                            for (const link of links) {
                                const text = link.innerText.trim();
                                const href = link.href || '';
                                const isYazhiLink = href.includes('/fenxi/yazhi-') || href.includes('odds.500.com/fenxi/yazhi-') || href.includes('yazhi-');
                                const isOuzhiLink = href.includes('/fenxi/ouzhi-') || href.includes('odds.500.com/fenxi/ouzhi-') || href.includes('ouzhi-');

                                if (!analysis_url && (isYazhiLink || (text === '亚' && href.includes('odds.500.com')))) {
                                    analysis_url = href;
                                }
                                if (!euro_odds_url && (isOuzhiLink || (text === '欧' && href.includes('odds.500.com')))) {
                                    euro_odds_url = href;
                                }
                                if (analysis_url && euro_odds_url) break;
                            }
                            if (analysis_url && euro_odds_url) break;
                        }
                        if (!analysis_url && euro_odds_url && euro_odds_url.includes('ouzhi') && euro_odds_url.includes('odds.500.com')) {
                            analysis_url = euro_odds_url.replace('ouzhi', 'yazhi');
                        }

                        matches.push({
                            match_id,
                            league,
                            round,
                            match_time,
                            status,
                            home_team,
                            home_rank,
                            handicap,
                            away_team,
                            away_rank,
                            halftime_score,
                            win_odds,
                            let_odds,
                            avg_euro,
                            william_odds,
                            aust_odds,
                            bet365_odds,
                            royal_odds,
                            analysis_url,
                            euro_odds_url
                        });

                    } catch (err) {
                        console.error(`处理第 ${i} 行时出错:`, err.message);
                    }
                }

                console.log(`成功提取 ${matches.length} 场比赛数据`);
                return matches;
            }
            """

            # Execute extraction and get results
            extracted_data = await page.evaluate(extraction_script)

            # Convert to MatchData objects
            matches = []
            for item in extracted_data:
                match = MatchData(
                    match_id=item.get('match_id', ''),
                    league=item.get('league', ''),
                    round=item.get('round', ''),
                    match_time=item.get('match_time', ''),
                    status=item.get('status', ''),
                    home_team=item.get('home_team', ''),
                    home_rank=item.get('home_rank', ''),
                    handicap=item.get('handicap', ''),
                    away_team=item.get('away_team', ''),
                    away_rank=item.get('away_rank', ''),
                    halftime_score=item.get('halftime_score', ''),
                    win_odds=item.get('win_odds', ''),
                    let_odds=item.get('let_odds', ''),
                    avg_euro=item.get('avg_euro', ''),
                    william_odds=item.get('william_odds', ''),
                    aust_odds=item.get('aust_odds', ''),
                    bet365_odds=item.get('bet365_odds', ''),
                    royal_odds=item.get('royal_odds', ''),
                    analysis_url=item.get('analysis_url', ''),
                    euro_odds_url=item.get('euro_odds_url', '')
                )
                matches.append(match)

            # Print table to terminal with detailed odds
            print("\n" + "=" * 200)
            print(f"{'场次':<8} {'赛事':<8} {'轮次':<12} {'时间':<14} {'状态':<4} {'主队':<22} {'让球':<12} {'客队':<22} {'胜负':<12} {'让球':<12} {'均欧':<10}")
            print("=" * 200)

            for match in matches:
                home_rank_str = f"[{match.home_rank}]" if match.home_rank else ""
                away_rank_str = f"[{match.away_rank}]" if match.away_rank else ""

                print(f"{match.match_id:<8} {match.league:<8} {match.round:<12} {match.match_time:<14} "
                      f"{match.status:<4} {home_rank_str}{match.home_team:<22} {match.handicap:<12} "
                      f"{match.away_team:<22}{away_rank_str} {match.win_odds:<12} {match.let_odds:<12} {match.avg_euro:<10}")

            print("=" * 160)
            print(f"\n✓ 成功从浏览器提取 {len(matches)} 场比赛数据")

            await browser.close()
            return matches

    except PlaywrightError as e:
        global LAST_SCRAPE_ERROR
        import traceback
        LAST_SCRAPE_ERROR = traceback.format_exc().strip()
        print(f"Playwright 错误: {e}")
        print("回退到生成示例数据...")
        fallback_data = _generate_fallback_matches()
        for match in fallback_data:
            match.is_fallback = True
        return fallback_data
    except Exception as e:
        global LAST_SCRAPE_ERROR
        import traceback
        LAST_SCRAPE_ERROR = traceback.format_exc().strip()
        print(f"浏览器自动化失败: {e}")
        print("回退到生成示例数据...")
        fallback_data = _generate_fallback_matches()
        for match in fallback_data:
            match.is_fallback = True
        return fallback_data


def _generate_fallback_matches(count: int = 15) -> List[MatchData]:
    """Generate fallback match data when browser scraping fails."""
    matches = []
    for i in range(1, count + 1):
        match = MatchData(
            match_id=f"周一{i:03d}",
            league=f"竞彩联赛",
            round=f"第{i}轮",
            match_time=f"01-{i+5:02d} 14:00",
            status="未",
            home_team=f"主队{i}",
            home_rank=f"{i:02d}",
            handicap=f"受平手/半球",
            away_team=f"客队{i}",
            away_rank=f"{20-i:02d}",
            halftime_score="-",
            win_odds="",
            let_odds="",
            avg_euro="",
            william_odds="",
            aust_odds="",
            bet365_odds="",
            royal_odds=""
        )
        matches.append(match)
    print(f"生成了 {len(matches)} 条示例数据")
    return matches


async def fetch_asian_handicap_data(
    matches: List[MatchData],
    headless: bool = True,
    timeout: int = 30000
) -> List[MatchData]:
    """
    Fetch Asian handicap analysis data for each match by visiting their analysis pages.

    This function navigates to each match's Asian handicap analysis page (via the "亚" link)
    and extracts the Crown (冠) row data from the Initial Handicap (初始盘口) column.

    Args:
        matches: List of MatchData objects with analysis_url populated
        headless: Run browser in headless mode
        timeout: Page load timeout in milliseconds

    Returns:
        Updated MatchData objects with asian_handicap, home_water, away_water populated
    """
    from playwright.async_api import async_playwright, Error as PlaywrightError

    print("\n" + "=" * 80)
    print("正在获取亚洲盘口分析数据...")
    print("=" * 80)

    # Filter matches that have analysis URLs
    matches_with_urls = [(i, m) for i, m in enumerate(matches) if m.analysis_url]

    if not matches_with_urls:
        print("⚠ 未找到任何分析页面链接")
        return matches

    print(f"找到 {len(matches_with_urls)} 个分析页面链接")

    # Debug: Print the analysis URLs
    for i, m in matches_with_urls[:3]:  # Print first 3
        print(f"  - {m.match_id}: {m.analysis_url}")

    try:
        async with async_playwright() as p:
            # Launch browser with anti-detection settings
            browser = await p.chromium.launch(
                headless=headless,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox',
                ]
            )

            # Create context with realistic user agent
            context = await browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                viewport={'width': 1920, 'height': 1080},
                locale='zh-CN',
            )

            extraction_script = """
            () => {
                const targetTable = document.getElementById('datatb');
                if (!targetTable) {
                    return { handicap: '', homeWater: '', awayWater: '' };
                }

                const rows = targetTable.querySelectorAll('tbody tr');

                for (const row of rows) {
                    const cells = row.querySelectorAll('td');
                    if (cells.length < 5) continue;

                    // Column 2 (index 1) contains bookmaker name
                    const bookmakerCell = cells[1];
                    const bookmakerLink = bookmakerCell.querySelector('a');
                    let bookmakerName = '';
                    let bookmakerTitle = '';

                    if (bookmakerLink) {
                        bookmakerName = bookmakerLink.innerText.trim();
                        bookmakerTitle = bookmakerLink.title || bookmakerLink.getAttribute('title') || '';
                    } else {
                        bookmakerName = bookmakerCell.innerText.trim();
                    }

                    // Check if this is the Crown bookmaker (冠 with optional asterisk)
                    if (bookmakerName.includes('冠') || bookmakerTitle.includes('冠')) {
                        // Search through cells to find nested tables with handicap data
                        // The INITIAL handicap table typically has cells WITHOUT arrow indicators
                        for (let i = 0; i < cells.length; i++) {
                            const nestedTable = cells[i].querySelector('table.pl_table_data');
                            if (nestedTable) {
                                const nestedRow = nestedTable.querySelector('tr');
                                if (nestedRow) {
                                    const nestedCells = nestedRow.querySelectorAll('td');
                                    if (nestedCells.length >= 3) {
                                        const cell0 = nestedCells[0].innerText.trim();
                                        const cell2 = nestedCells[2].innerText.trim();

                                        // Check if cells DON'T have arrow indicators (↑/↓/升/降)
                                        // Initial handicap data is plain numbers without trends
                                        const hasArrows = cell0.includes('↑') || cell0.includes('↓') ||
                                                         cell2.includes('↑') || cell2.includes('↓') ||
                                                         cell0.includes('升') || cell0.includes('降') ||
                                                         cell2.includes('升') || cell2.includes('降');

                                        // Check if this looks like handicap data (decimal numbers)
                                        const looksLikeOdds = /^[0-9.]+$/.test(cell0.replace('↑', '').replace('↓', '')) &&
                                                              /^[0-9.]+$/.test(cell2.replace('↑', '').replace('↓', ''));

                                        if (looksLikeOdds && !hasArrows) {
                                            // This is likely the INITIAL handicap data (no arrows)
                                            const homeWater = cell0;
                                            const handicap = nestedCells[1].innerText.trim();
                                            const awayWater = cell2;
                                            return { handicap, homeWater, awayWater };
                                        }
                                    }
                                }
                            }
                        }

                        // If no table without arrows found, fall back to first valid table
                        for (let i = 0; i < cells.length; i++) {
                            const nestedTable = cells[i].querySelector('table.pl_table_data');
                            if (nestedTable) {
                                const nestedRow = nestedTable.querySelector('tr');
                                if (nestedRow) {
                                    const nestedCells = nestedRow.querySelectorAll('td');
                                    if (nestedCells.length >= 3) {
                                        const homeWater = nestedCells[0].innerText.trim();
                                        const handicap = nestedCells[1].innerText.trim();
                                        const awayWater = nestedCells[2].innerText.trim();
                                        return { handicap, homeWater, awayWater };
                                    }
                                }
                            }
                        }

                        return { handicap: '', homeWater: '', awayWater: '' };
                    }
                }

                return { handicap: '', homeWater: '', awayWater: '' };
            }
            """

            max_concurrency = 4
            semaphore = asyncio.Semaphore(max_concurrency)

            async def fetch_one(idx, match_index, match):
                async with semaphore:
                    page = await context.new_page()
                    try:
                        print(f"[{idx}/{len(matches_with_urls)}] 正在获取 {match.match_id} 的亚洲盘口数据...")
                        print(f"  URL: {match.analysis_url}")

                        # Navigate to the analysis page
                        await page.goto(match.analysis_url, wait_until='domcontentloaded', timeout=timeout)

                        # Wait for content to load
                        await asyncio.sleep(5)

                        # Scroll to bottom to ensure all content is loaded (some sites lazy-load)
                        await page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                        await asyncio.sleep(2)

                        # Scroll back to top
                        await page.evaluate('window.scrollTo(0, 0)')
                        await asyncio.sleep(2)

                        # Try to wait for specific elements that might contain the odds data
                        try:
                            await page.wait_for_selector('table', timeout=5000)
                        except Exception:
                            pass

                        # Additional wait for dynamic content
                        await asyncio.sleep(2)

                        # Get page title for debugging
                        title = await page.title()
                        print(f"  页面标题: {title}")

                        result = await page.evaluate(extraction_script)

                        # Update the match data
                        matches[match_index].asian_handicap = result.get('handicap', '')
                        matches[match_index].home_water = result.get('homeWater', '')
                        matches[match_index].away_water = result.get('awayWater', '')

                        print(f"  ✓ 盘: {matches[match_index].asian_handicap or 'N/A'}, "
                              f"水1: {matches[match_index].home_water or 'N/A'}, "
                              f"水2: {matches[match_index].away_water or 'N/A'}")

                        # Small delay to be respectful to the server
                        await asyncio.sleep(0.5)

                    except PlaywrightError as e:
                        print(f"  ✗ 无法获取 {match.match_id} 的数据: {e}")
                    except Exception as e:
                        print(f"  ✗ 处理 {match.match_id} 时出错: {e}")
                    finally:
                        await page.close()

            tasks = [
                fetch_one(idx, match_index, match)
                for idx, (match_index, match) in enumerate(matches_with_urls, 1)
            ]
            await asyncio.gather(*tasks)

            await browser.close()

    except PlaywrightError as e:
        print(f"浏览器自动化失败: {e}")
    except Exception as e:
        print(f"获取亚洲盘口数据时出错: {e}")

    print("=" * 80)
    print(f"✓ 亚洲盘口数据获取完成")
    print("=" * 80)

    return matches


def _normalize_odds_url(url: str) -> str:
    """Normalize odds URL to absolute https URL."""
    if not url:
        return ""
    url = url.strip()
    if url.startswith("//"):
        return f"https:{url}"
    if url.startswith("/"):
        return f"https://odds.500.com{url}"
    return url


def _derive_ouzhi_url(analysis_url: str, euro_odds_url: str = "") -> str:
    """Derive the 百家欧赔 URL from the 亚盘对比 URL."""
    if euro_odds_url:
        return _normalize_odds_url(euro_odds_url)

    if not analysis_url:
        return ""

    url = _normalize_odds_url(analysis_url)
    if "odds.500.com" not in url and "/fenxi/" not in url and "yazhi-" not in url:
        return ""

    if "ouzhi" in url:
        return url
    if "yazhi" in url:
        return url.replace("yazhi", "ouzhi")

    match = re.search(r"(\d{5,})", url)
    if match:
        return f"https://odds.500.com/fenxi/ouzhi-{match.group(1)}.shtml"

    return ""


async def fetch_euro_kelly_data(
    matches: List[MatchData],
    headless: bool = True,
    timeout: int = 30000
) -> List[MatchData]:
    """
    Fetch European odds Kelly index (百家欧赔) for Crown (冠) row.

    Extracts the "即时凯利" values for 胜/负 from the Crown row and stores them in
    MatchData.euro_kelly_win / MatchData.euro_kelly_lose.
    """
    from playwright.async_api import async_playwright, Error as PlaywrightError

    print("\n" + "=" * 80)
    print("正在获取百家欧赔即时凯利数据...")
    print("=" * 80)

    matches_with_urls = []
    for i, match in enumerate(matches):
        ouzhi_url = _derive_ouzhi_url(match.analysis_url, match.euro_odds_url)
        if ouzhi_url:
            matches_with_urls.append((i, match, ouzhi_url))

    if not matches_with_urls:
        print("⚠ 未找到任何百家欧赔页面链接")
        return matches

    print(f"找到 {len(matches_with_urls)} 个百家欧赔页面链接")
    for i, m, ouzhi_url in matches_with_urls[:3]:
        print(f"  - {m.match_id}: {ouzhi_url}")

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=headless,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox',
                ]
            )

            context = await browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                viewport={'width': 1920, 'height': 1080},
                locale='zh-CN',
            )

            extraction_function = """
            (rows) => {
                for (const row of rows) {
                    const cells = row.querySelectorAll('td');
                    if (cells.length < 6) continue;

                    const bookmakerCell = cells[1];
                    const bookmakerLink = bookmakerCell.querySelector('a');
                    let bookmakerName = '';
                    let bookmakerTitle = '';

                    if (bookmakerLink) {
                        bookmakerName = (bookmakerLink.textContent || '').trim();
                        bookmakerTitle = bookmakerLink.title || bookmakerLink.getAttribute('title') || '';
                    } else {
                        bookmakerName = (bookmakerCell.textContent || '').trim();
                        bookmakerTitle = bookmakerCell.title || bookmakerCell.getAttribute('title') || '';
                    }

                    if (bookmakerName.includes('冠') || bookmakerTitle.includes('冠')) {
                        let kellyTable = null;
                        if (cells[5]) {
                            kellyTable = cells[5].querySelector('table.pl_table_data');
                        }

                        if (!kellyTable) {
                            const tables = Array.from(row.querySelectorAll('table.pl_table_data'));
                            const candidates = tables.filter(table => {
                                const firstRow = table.querySelector('tr');
                                if (!firstRow) return false;
                                return firstRow.querySelectorAll('td').length >= 3;
                            });
                            if (candidates.length) {
                                kellyTable = candidates[candidates.length - 1];
                            }
                        }

                        if (!kellyTable) {
                            return { win: '', lose: '' };
                        }

                        const rows = Array.from(kellyTable.querySelectorAll('tr'));
                        if (!rows.length) {
                            return { win: '', lose: '', win2: '', lose2: '' };
                        }

                        const extractRow = (row) => {
                            if (!row) return { win: '', draw: '', lose: '' };
                            const tds = row.querySelectorAll('td');
                            if (tds.length >= 3) {
                                return {
                                    win: (tds[0].textContent || '').trim(),
                                    draw: (tds[1].textContent || '').trim(),
                                    lose: (tds[2].textContent || '').trim()
                                };
                            }
                            return { win: '', draw: '', lose: '' };
                        };

                        const isNumeric = (value) => /^\\d+(\\.\\d+)?$/.test(value);
                        const candidates = [];
                        for (const row of rows) {
                            const extracted = extractRow(row);
                            if (isNumeric(extracted.win) && isNumeric(extracted.draw) && isNumeric(extracted.lose)) {
                                candidates.push(extracted);
                            }
                        }

                        const first = candidates[0] || extractRow(kellyTable.querySelector('tr.td_show_cp')) || { win: '', draw: '', lose: '' };
                        const second = candidates.length > 1 ? candidates[1] : { win: '', draw: '', lose: '' };

                        return { win: first.win, draw: first.draw, lose: first.lose, win2: second.win, draw2: second.draw, lose2: second.lose };
                    }
                }

                return { win: '', draw: '', lose: '', win2: '', draw2: '', lose2: '' };
            }
            """

            extraction_on_page = f"""
            () => {{
                const extract = {extraction_function};
                const rows = document.querySelectorAll('#datatb tbody tr');
                return extract(rows);
            }}
            """

            extraction_from_html = f"""
            async (payload) => {{
                const params = new URLSearchParams({{
                    id: payload.id,
                    ctype: payload.ctype,
                    start: '0',
                    r: '1',
                    style: '0',
                    guojia: '0',
                    chupan: '1',
                    currentIndex: '0'
                }});
                const resp = await fetch(`/fenxi1/ouzhi.php?${{params.toString()}}`, {{ credentials: 'include' }});
                if (!resp.ok) {{
                    return {{ win: '', lose: '' }};
                }}
                const html = await resp.text();
                const doc = new DOMParser().parseFromString(html, 'text/html');
                const rows = doc.querySelectorAll('tr');
                const extract = {extraction_function};
                return extract(rows);
            }}
            """

            max_concurrency = 4
            semaphore = asyncio.Semaphore(max_concurrency)

            async def fetch_one(idx, match_index, match, ouzhi_url):
                async with semaphore:
                    page = await context.new_page()
                    try:
                        print(f"[{idx}/{len(matches_with_urls)}] 正在获取 {match.match_id} 的百家欧赔数据...")
                        print(f"  URL: {ouzhi_url}")

                        await page.goto(ouzhi_url, wait_until='domcontentloaded', timeout=timeout)

                        try:
                            await page.wait_for_selector('#datatb', timeout=8000)
                            await page.wait_for_function(
                                "() => document.querySelectorAll('#datatb tbody tr').length > 0",
                                timeout=8000
                            )
                            await page.wait_for_function(
                                "() => Array.from(document.querySelectorAll('#datatb tbody tr td:nth-child(2)')).some(td => td.innerText.includes('冠'))",
                                timeout=8000
                            )
                        except Exception:
                            pass

                        await asyncio.sleep(2)

                        result = await page.evaluate(extraction_on_page)

                        def _looks_numeric(value: str) -> bool:
                            return bool(value) and re.fullmatch(r"[0-9.]+", value.strip()) is not None

                        if not (_looks_numeric(result.get('win', '')) and _looks_numeric(result.get('lose', ''))):
                            match_id = re.search(r"ouzhi-(\\d+)\\.shtml", ouzhi_url)
                            fixture_id = match_id.group(1) if match_id else ""
                            if fixture_id:
                                fallback_result = await page.evaluate(
                                    extraction_from_html,
                                    {"id": fixture_id, "ctype": 1}
                                )
                                result = fallback_result

                        matches[match_index].euro_kelly_win = result.get('win', '')
                        matches[match_index].euro_kelly_draw = result.get('draw', '')
                        matches[match_index].euro_kelly_lose = result.get('lose', '')
                        matches[match_index].euro_kelly_win_2 = result.get('win2', '')
                        matches[match_index].euro_kelly_draw_2 = result.get('draw2', '')
                        matches[match_index].euro_kelly_lose_2 = result.get('lose2', '')

                        print(f"  ✓ 凯利胜: {matches[match_index].euro_kelly_win or 'N/A'}, "
                              f"凯利平: {matches[match_index].euro_kelly_draw or 'N/A'}, "
                              f"凯利负: {matches[match_index].euro_kelly_lose or 'N/A'}, "
                              f"凯利胜2: {matches[match_index].euro_kelly_win_2 or 'N/A'}, "
                              f"凯利平2: {matches[match_index].euro_kelly_draw_2 or 'N/A'}, "
                              f"凯利负2: {matches[match_index].euro_kelly_lose_2 or 'N/A'}")

                        await asyncio.sleep(0.5)

                    except PlaywrightError as e:
                        print(f"  ✗ 无法获取 {match.match_id} 的数据: {e}")
                    except Exception as e:
                        print(f"  ✗ 处理 {match.match_id} 时出错: {e}")
                    finally:
                        await page.close()

            tasks = [
                fetch_one(idx, match_index, match, ouzhi_url)
                for idx, (match_index, match, ouzhi_url) in enumerate(matches_with_urls, 1)
            ]
            await asyncio.gather(*tasks)

            await browser.close()

    except PlaywrightError as e:
        print(f"浏览器自动化失败: {e}")
    except Exception as e:
        print(f"获取百家欧赔数据时出错: {e}")

    print("=" * 80)
    print("✓ 百家欧赔即时凯利数据获取完成")
    print("=" * 80)

    return matches


async def fetch_enhanced_odds_data(match_ids: List[str]) -> Dict[str, Dict]:
    """
    Fetch enhanced odds data for specific matches using browser automation.
    This can navigate to detailed odds pages for each match.

    Args:
        match_ids: List of match IDs to fetch odds for

    Returns:
        Dictionary mapping match_id to odds data
    """
    from playwright.async_api import async_playwright

    odds_data = {}

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()

            for match_id in match_ids:
                try:
                    # Construct detailed odds URL for the match
                    detail_url = f"{ODDS_URL}jczq/{match_id.replace('周', '')}.shtml"
                    await page.goto(detail_url, timeout=15000)

                    # Extract odds data
                    odds_script = """
                    () => {
                        return {
                            kelly: document.querySelector('.kelly-value')?.innerText,
                            asian_handicap: document.querySelector('.ah-value')?.innerText,
                            over_under: document.querySelector('.ou-value')?.innerText,
                            history: Array.from(document.querySelectorAll('.odds-history tr')).map(tr => ({
                                time: tr.cells[0]?.innerText,
                                home: tr.cells[1]?.innerText,
                                draw: tr.cells[2]?.innerText,
                                away: tr.cells[3]?.innerText
                            }))
                        };
                    }
                    """

                    match_odds = await page.evaluate(odds_script)
                    odds_data[match_id] = match_odds

                    # Small delay to be respectful to the server
                    await asyncio.sleep(0.5)

                except Exception as e:
                    print(f"无法获取比赛 {match_id} 的赔率数据: {e}")
                    odds_data[match_id] = {}

            await browser.close()

    except Exception as e:
        print(f"增强赔率数据获取失败: {e}")

    return odds_data


# ============================================================================
# EXCEL GENERATION
# ============================================================================


def create_template_workbook():
    """Create a new workbook with headers and basic formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "足球彩票分析模板"
    return wb, ws


def set_column_widths(ws):
    """Set appropriate column widths."""
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width


def merge_header_cells(ws):
    """Merge cells for main section headers."""
    for start_cell, end_cell in MERGED_CELLS:
        ws.merge_cells(f"{start_cell}:{end_cell}")


def style_header_rows(ws):
    """Apply styling to header rows."""
    # Style Row 1
    for col_idx, header in enumerate(MAIN_HEADERS, start=1):
        if header:
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER_STYLE

    # Style Row 2
    for col_idx, header in enumerate(SUB_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = BORDER_STYLE

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20


def add_match_data(ws, start_row: int, match: MatchData) -> int:
    """
    Add comprehensive match data to the worksheet.

    Args:
        ws: Worksheet object
        start_row: Starting row number
        match: MatchData object with comprehensive information

    Returns:
        Number of rows added
    """
    current_row = start_row

    # Match ID row
    ws.cell(row=current_row, column=1, value=match.match_id)
    ws.cell(row=current_row, column=2, value=match.handicap)

    # Add Asian handicap data if available (replaces the odds data)
    if match.asian_handicap:
        ws.cell(row=current_row, column=2, value=match.asian_handicap)  # 盘
    if match.home_water:
        ws.cell(row=current_row, column=3, value=match.home_water)  # 水1
    if match.away_water:
        ws.cell(row=current_row, column=4, value=match.away_water)  # 水2

    # Add odds data if available (only if Asian handicap data is not present)
    if not match.asian_handicap and match.win_odds:
        ws.cell(row=current_row, column=3, value=match.win_odds)
    if not match.asian_handicap and match.let_odds:
        ws.cell(row=current_row, column=4, value=match.let_odds)

    # Style match ID row
    for col_idx in range(1, 19):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.fill = MATCH_ID_FILL
        cell.font = MATCH_ID_FONT
        cell.alignment = DATA_ALIGNMENT
        cell.border = BORDER_STYLE

    current_row += 1

    # Add time point rows
    for idx, time_label in enumerate(TIME_LABELS):
        ws.cell(row=current_row, column=1, value=time_label)
        if idx == 0:
            if match.euro_kelly_win:
                ws.cell(row=current_row, column=3, value=match.euro_kelly_win)
            if match.euro_kelly_lose:
                ws.cell(row=current_row, column=4, value=match.euro_kelly_lose)
            if match.euro_kelly_win_2:
                ws.cell(row=current_row, column=6, value=match.euro_kelly_win_2)
            if match.euro_kelly_lose_2:
                ws.cell(row=current_row, column=7, value=match.euro_kelly_lose_2)
            if match.euro_kelly_draw:
                ws.cell(row=current_row, column=9, value=match.euro_kelly_draw)
            if match.euro_kelly_draw_2:
                ws.cell(row=current_row, column=10, value=match.euro_kelly_draw_2)

        # Style the row
        for col_idx in range(1, 19):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.alignment = DATA_ALIGNMENT
            cell.border = BORDER_STYLE

            if col_idx == 1:
                cell.fill = TIME_LABEL_FILL
                cell.font = TIME_LABEL_FONT
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        current_row += 1

    return len(TIME_LABELS) + 1


def generate_browser_template(
    filename: str = "live_betting_template.xlsx",
    url: str = BASE_URL,
    headless: bool = True,
    max_matches: Optional[int] = None,
    fetch_enhanced_odds: bool = False,
    fetch_asian_handicap: bool = False
):
    """
    Generate Excel betting analysis template using browser automation.

    Args:
        filename: Output Excel filename
        url: URL to scrape match data from
        headless: Run browser in headless mode
        max_matches: Maximum number of matches to include
        fetch_enhanced_odds: Whether to fetch enhanced odds data (slower)
        fetch_asian_handicap: Whether to fetch Asian handicap analysis data
    """
    print("=" * 80)
    print("浏览器自动化 - 足球彩票分析模板生成器")
    print("=" * 80)

    # Fetch match data using browser automation
    matches = asyncio.run(fetch_matches_with_browser(url, headless=headless))

    if not matches:
        print("未找到比赛数据，生成空模板...")
        matches = _generate_fallback_matches(5)
    else:
        # Filter for the most recent day (based on the first match's weekday prefix)
        first_match_id = matches[0].match_id
        # Extract "周X" (first 2 chars)
        if len(first_match_id) >= 2 and first_match_id.startswith("周"):
            current_weekday = first_match_id[:2]
            original_count = len(matches)
            matches = [m for m in matches if m.match_id.startswith(current_weekday)]
            filtered_count = len(matches)
            print(f"筛选最近一天比赛 ({current_weekday}): 从 {original_count} 场保留至 {filtered_count} 场")

    # Limit matches if specified
    if max_matches is not None:
        matches = matches[:max_matches]

    # Optionally fetch enhanced odds data
    odds_data = {}
    if fetch_enhanced_odds:
        print("\n正在获取增强赔率数据...")
        match_ids = [m.match_id for m in matches]
        odds_data = asyncio.run(fetch_enhanced_odds_data(match_ids))

    # Optionally fetch Asian handicap analysis data
    if fetch_asian_handicap:
        matches = asyncio.run(fetch_asian_handicap_data(matches, headless=headless))
        matches = asyncio.run(fetch_euro_kelly_data(matches, headless=headless))

    # Create workbook and worksheet
    wb, ws = create_template_workbook()

    # Apply formatting
    set_column_widths(ws)
    merge_header_cells(ws)
    style_header_rows(ws)

    # Add match data
    current_row = 3
    for match in matches:
        rows_added = add_match_data(ws, current_row, match)
        current_row += rows_added

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Add metadata sheet
    info_sheet = wb.create_sheet("数据来源信息")
    info_sheet.column_dimensions['A'].width = 20
    info_sheet.column_dimensions['B'].width = 40

    info_data = [
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["数据来源", url],
        ["抓取方式", "浏览器自动化 (Playwright)"],
        ["比赛数量", len(matches)],
        ["增强赔率数据", "是" if fetch_enhanced_odds else "否"],
        ["亚洲盘口分析", "是" if fetch_asian_handicap else "否"],
        ["百家欧赔即时凯利", "是" if fetch_asian_handicap else "否"],
    ]

    for row_idx, (label, value) in enumerate(info_data, start=1):
        info_sheet.cell(row=row_idx, column=1, value=label)
        info_sheet.cell(row=row_idx, column=2, value=value)

    # Save workbook
    wb.save(filename)

    print("=" * 80)
    print(f"✓ Excel 模板 '{filename}' 生成成功！")
    print(f"  - 包含 {len(matches)} 场真实比赛数据")
    print(f"  - 使用浏览器自动化抓取数据")
    if fetch_enhanced_odds:
        print(f"  - 已包含增强赔率数据")
    print("=" * 80)


def main():
    """Entry point for the CLI command."""
    import argparse

    parser = argparse.ArgumentParser(
        description="使用浏览器自动化生成足球彩票分析 Excel 模板"
    )
    parser.add_argument(
        "-o", "--output",
        default="live_betting_template.xlsx",
        help="输出 Excel 文件名 (默认: live_betting_template.xlsx)"
    )
    parser.add_argument(
        "-u", "--url",
        default=BASE_URL,
        help=f"要抓取的 URL (默认: {BASE_URL})"
    )
    parser.add_argument(
        "--no-headless",
        action="store_true",
        help="显示浏览器窗口（用于调试）"
    )
    parser.add_argument(
        "-m", "--max-matches",
        type=int,
        default=None,
        help="最多包含的比赛数量"
    )
    parser.add_argument(
        "--enhanced-odds",
        action="store_true",
        help="获取增强赔率数据（更慢但更全面）"
    )
    parser.add_argument(
        "--asian-handicap",
        action="store_true",
        help="获取亚洲盘口分析数据（点击亚盘分析页面）"
    )

    args = parser.parse_args()

    generate_browser_template(
        filename=args.output,
        url=args.url,
        headless=not args.no_headless,
        max_matches=args.max_matches,
        fetch_enhanced_odds=args.enhanced_odds,
        fetch_asian_handicap=args.asian_handicap
    )


if __name__ == "__main__":
    main()
